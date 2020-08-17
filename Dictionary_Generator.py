#!usr/bin/env python3

from openpyxl import load_workbook
from SupportFunctions import *
import lifelines

workbook = load_workbook(filename="Dataset.xlsx")
CDR3File = workbook['metastatic and primary CDR3s b']
DNAH9 = workbook['DNAH9']
Clinical = workbook['clinical']
PatientSurvivalChart = workbook['PatientSurvivalChart']

'''
Dictionary Example:
    {
        "TCGA-D3-A1Q4":
        {
            "Patient ID": TCGA-D3-A1Q4, "TRA CDR3":[abc,abc],
            "TRB CDR3":[abc,abc], "Mutations": ["S1234D"], "Months Left": 10,
            "Complimentary": True
        }
    }

'''

def main(Pathway,IncludeAlive):
    Major_Dictionary = {}
    
    for x in range(2, DNAH9.max_row + 1):
        Patient_ID = DNAH9["A{cellRow}".format(cellRow = x)].value
        
        # If patient ID already in dictionary
        # simply append mutations array to include the new one
        if Patient_ID in Major_Dictionary:
            Mutation = DNAH9["E{cellRow}".format(cellRow = x)].value
            Major_Dictionary[Patient_ID]["Mutations"].append(Mutation)
        
        else:
            MonthsLeft = MonthsLeftFunc(Patient_ID)
            if IncludeAlive == False:
                if MonthsLeft != "N/A":
                    Major_Dictionary[Patient_ID] = {"Patient ID":Patient_ID,"TRA CDR3":[], "TRB CDR3":[],
                    "Mutations": [], "Months Left": False, "Complimentary": False,
                    "Method":"new"}
                    
                    Mutation = DNAH9["E{cellRow}".format(cellRow = x)].value
                    Major_Dictionary[Patient_ID]["Mutations"].append(Mutation)

                    Major_Dictionary[Patient_ID]["Months Left"] = MonthsLeft
                    
            elif IncludeAlive == True:
                Major_Dictionary[Patient_ID] = {"Patient ID":Patient_ID,"TRA CDR3":[], "TRB CDR3":[],
                "Mutations": [], "Months Left": False, "Complimentary": False,
                "Method":"new"}
                
                Mutation = DNAH9["E{cellRow}".format(cellRow = x)].value
                Major_Dictionary[Patient_ID]["Mutations"].append(Mutation)
                
                Major_Dictionary[Patient_ID]["Months Left"] = MonthsLeft

    # For each CDR3 value add it to the patient it belongs to
    # add CDR3 domain to TRA or TRB array depending on receptor
    for x in range(2, CDR3File.max_row + 1):
        Patient_ID = CDR3File["A{cellRow}".format(cellRow = x)].value
        
        if Patient_ID in Major_Dictionary:
            CDR3 = CDR3File["B{cellRow}".format(cellRow = x)].value
            Receptor = CDR3File["C{cellRow}".format(cellRow = x)].value

            Major_Dictionary[Patient_ID]["{receptor} CDR3".format(receptor = Receptor)].append(CDR3)
            
    for Patient_ID in Major_Dictionary:
        for Mutation in Major_Dictionary[Patient_ID]["Mutations"]:
            
            ######
            #    Marks all Mutations with peptides as both so calculations
            #    occur with both TRA and TRB CDR3's
            if Pathway == "Old":
                MutationSide = "Both"
            #
            ######
                    
            else:   # Computes Left or Right for new method
                MutationSide = LeftorRight(Mutation[1:len(Mutation)-1])

            ######
            #   New Method Calculations...For each mutation we calculate
            #   complimentary with all CDR3's in TRA or TRB depending on side
            #   Marks done as True if one complimetary mutation and CDR3
            #   combo happens
            #
            if MutationSide == "Left":
                for CDR3 in Major_Dictionary[Patient_ID]["TRA CDR3"]:
                    if ComplimentaryFunction(CDR3,Mutation) == True and Major_Dictionary[Patient_ID]["Complimentary"] == False:
                        Major_Dictionary[Patient_ID]["Complimentary"] = True
            elif MutationSide == "Right":
                for CDR3 in Major_Dictionary[Patient_ID]["TRB CDR3"]:
                    if ComplimentaryFunction(CDR3,Mutation) == True and Major_Dictionary[Patient_ID]["Complimentary"] == False:
                        Major_Dictionary[Patient_ID]["Complimentary"] = True
            elif MutationSide == "Both":
                for CDR3 in Major_Dictionary[Patient_ID]["TRA CDR3"]:
                    if ComplimentaryFunction(CDR3,Mutation) == True and Major_Dictionary[Patient_ID]["Complimentary"] == False:
                        Major_Dictionary[Patient_ID]["Complimentary"] = True
                for CDR3 in Major_Dictionary[Patient_ID]["TRB CDR3"]:
                    if ComplimentaryFunction(CDR3,Mutation) == True and Major_Dictionary[Patient_ID]["Complimentary"] == False:
                        Major_Dictionary[Patient_ID]["Complimentary"] = True
            elif MutationSide == "N/A":
                #print("No Peptides", Patient_ID, Mutation)
                _=0
            #
            ######

    PercentAlive(Major_Dictionary,Pathway)
    ExcelOutput(Major_Dictionary,Pathway, True)
    LogRankTest(Major_Dictionary,Pathway)

    return Major_Dictionary

'''Code to run main function with Alive included''' 
Major_Dictionary_Old = main("Old",True)
Major_Dictionary = main("New",True)

for Patient_ID in Major_Dictionary_Old:
    ID_1 = Major_Dictionary_Old[Patient_ID]["Patient ID"]
    for Patient_ID in Major_Dictionary:
        ID_2 = Major_Dictionary[Patient_ID]["Patient ID"]
        if ID_1 == ID_2:
            if Major_Dictionary_Old[Patient_ID]["Complimentary"] == True:
                if Major_Dictionary[Patient_ID]["Complimentary"] == False:
                    print(ID_1)





