#!usr/bin/env python3
#Note: 1 Month = 30 days

from openpyxl import load_workbook

AliveStillComp = {}
AliveStillNon = {}

x = 0
while x < 501:
    AliveStillComp[x] = 0
    AliveStillNon[x] = 0
    x+=1

workbook = load_workbook(filename="EditedDataset.xlsx")
sheet = workbook['metastatic and primary CDR3s b']
sheet4 = workbook['PatientSurvivalChart']

def PercentAlive(sheet,sheet4):
    x = 2
    while x < sheet.max_row:
        Patient_ID = sheet["A{cellRow}".format(cellRow = x)].value
        MonthsLeft = sheet["I{cellRow}".format(cellRow = x)].value
        Complimentary = sheet["J{cellRow}".format(cellRow = x)].value
        while Patient_ID == sheet["A{cellRow}".format(cellRow = x)].value:
            x+=1
        if MonthsLeft != "'--":
            if Complimentary == True:
                AliveStillComp[MonthsLeft] = AliveStillComp[MonthsLeft] + 1
            else:
                AliveStillNon[MonthsLeft] = AliveStillNon[MonthsLeft] + 1
    x = 3
    CompCurrent = 315
    NonCompCurrent = 315

    while x <= 500:
        CompCurrent -= AliveStillComp[x]
        NonCompCurrent -= AliveStillNon[x]
        sheet4["B{cellRow}".format(cellRow = x)] = CompCurrent/315*100
        sheet4["D{cellRow}".format(cellRow = x)] = NonCompCurrent/315*100
        x += 1

PercentAlive(sheet,sheet4)

workbook.save(filename="EditedDataset3.xlsx")
