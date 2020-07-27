from openpyxl import load_workbook

workbook = load_workbook(filename="Dataset.xlsx")
sheet = workbook['metastatic and primary CDR3s b']
sheet2 = workbook['DNAH9']
sheet3 = workbook['clinical']
sheet4 = workbook['PatientSurvivalChart']

def get_KD_original():
    """ Function which returns the original KD hydropathy lookup table
    """    
    
    return  {'I': 4.5,'V': 4.2,'L': 3.8,'F': 2.8,
             'C': 2.5,'M': 1.9,'A': 1.8,'G': -0.4,
             'T': -0.7,'S': -0.8,'W': -0.9,'Y': -1.3,
             'P': -1.6,'H': -3.2,'E': -3.5,'Q': -3.5,
             'D': -3.5,'N': -3.5,'K': -3.9,'R': -4.5}

def get_AA_charge(AA):
    """ Function which returns the Charge Value for an Amino Acid
    """    
    AA_Charge_Table = {'I': 0,'V': 0,'L': 0,'F': 0,
             'C': 0,'M': 0,'A': 0,'G': 0,'T': 0,
             'S': 0,'W': 0,'Y': 0,'P': 0,'H': .1,
             'E': -1,'Q': 0,'D': -1,'N': 0,'K': 1,
             'R': 1}
    return AA_Charge_Table[AA]

def CalculateNCPR(Sequence):
    """ Function which Calculates NCPR by adding all the charges of each AA
        and divides by the total number of AAs in the Sequence
    """   
    totalCharge = 0
    for AA in Sequence:
        Charge = get_AA_charge(AA)
        totalCharge += Charge
    NCPR = totalCharge/len(Sequence)
    return NCPR

def CheckID(ID,sheet2):
    """ Checks ID from "metastatic and primary CDR3s b" sheet
    to find match in "DNAH9". If match is found returns Reference_Allele,
    Tumor_Seq_Allele1, Tumor_Seq_Allele2
    """ 
    for y in range(2, sheet2.max_row + 1):
        ID2 = sheet2["A{cellRow}".format(cellRow = y)].value
        if ID == ID2:
            AminoAcids = sheet2["C{cellRow}".format(cellRow = y)].value
            Reference_Allele = AminoAcids[0]
            Tumor_Seq_Allele = AminoAcids[2]
            return Reference_Allele, Tumor_Seq_Allele
    return "N/A", "N/A"

def MonthsLeftFunc(ID,sheet3):
    """Function to find the number of days a patient survived
    and return it in months. (One month equals 30 days)
    If patient is still alive it returns N/A
    """
    for y in range(2, sheet3.max_row + 1):
        ID2 = sheet3["B{cellRow}".format(cellRow = y)].value
        if ID == ID2:
            daysLeft = sheet3["J{cellRow}".format(cellRow = y)].value
            try:
                return int(daysLeft/30)
            except:
                return daysLeft
    return "N/A"

def Calc_AA_Charge_Δ(Ref_AA,TAA):
    """Function to Calculates Change in Charge by checking the
    change in charge of the Tumor_Seq_Allele1(TAA1) and
    Tumor_Seq_Allele2(TAA2) against the Reference Allele(Ref_AA)
    """
    if Ref_AA == "N/A":
        return "N/A"
    if TAA != "-":
        AA_Charge_Δ = get_AA_charge(TAA)-get_AA_charge(Ref_AA)
    return AA_Charge_Δ

#If AA_Charge_Δ = 0 multiplying makes NCPR_CS 0, so if  AA_Charge_Δ = 0
#AA_Charge_Δ is exluded from calculations

def Calc_NCPR_CS(AA_Charge_Δ,NCPR,Patient_ID):
    """Function to calculate NCPR by 
    """
    '''A positive value denotes a complementary score'''
    if AA_Charge_Δ == "N/A":
        return "N/A"
    NCPR_CS = AA_Charge_Δ * NCPR * -1
    return NCPR_CS

def MakeTrue(Patient_ID,sheet):
    """Function makes all instances of the Patient ID in metastatic
    and primary CDR3s b True"""
    for x in range(2, sheet.max_row + 1):
        if Patient_ID == sheet["A{cellRow}".format(cellRow = x)].value:
            sheet["J{cellRow}".format(cellRow = x)] = True

def PercentAlive(sheet,sheet4):
    """Function to calculate the survival charts"""
    
    NumPatients = 196   #Number of Patients in DNAH9
    
    AliveStillComp = {} #Holds the number of patients that died on each month
                        #from 1-500 that are complimentary
    
    AliveStillNon = {}  #Holds the number of patients that died on each month
                        #from 1-500 that are NONcomplimentary
    x = 0
    while x < 501: #Populates both dictionaries with 0 for 1-500
        AliveStillComp[x] = 0
        AliveStillNon[x] = 0
        x+=1

    x = 2
    while x < sheet.max_row:
        Patient_ID = sheet["A{cellRow}".format(cellRow = x)].value
        MonthsLeft = sheet["I{cellRow}".format(cellRow = x)].value
        Complimentary = sheet["J{cellRow}".format(cellRow = x)].value
        while Patient_ID == sheet["A{cellRow}".format(cellRow = x)].value:
            x+=1 # moves to the last instance of this Patient ID
        if MonthsLeft != "'--":
            if Complimentary == True:
                AliveStillComp[MonthsLeft] = AliveStillComp[MonthsLeft] + 1
                #Adds this patient to the month they died in the dictionary
                #If they were complimentary
            elif Complimentary == False:
                AliveStillNon[MonthsLeft] = AliveStillNon[MonthsLeft] + 1
                #Adds this patient to the month they died in the dictionary
                #If they were NONcomplimentary

    CompCurrent = 0     #Holds the number of people currently dead
                        #that were Complimentary
    
    NonCompCurrent = 0  #Holds the number of people currently dead
                        #that were Complimentary

    x =0    # X simulates the month
    while x <= 500:
        CompPatientNumber = 64
        NONCompPatientNumber = 81
        #The number of patients that are Complimentary and NonComplimentary        
        
        CompCurrent += AliveStillComp[x] 
        NonCompCurrent += AliveStillNon[x]
        #Each month we add the number of people that died that
        #month to CompCurrent and NonCompCurrent
        
        sheet4["B{cellRow}".format(cellRow = x+3)] = (CompPatientNumber-CompCurrent)/CompPatientNumber*100
        sheet4["D{cellRow}".format(cellRow = x+3)] = (NONCompPatientNumber-NonCompCurrent)/NONCompPatientNumber*100
        #Creat percentage of people left and add to Excel File next to month Number

        x += 1
        #Move Forward a Month
