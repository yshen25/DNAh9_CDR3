#!usr/bin/env python3
#NOTE: Amino Acid = AA

from openpyxl import load_workbook
from SupportFunctions import *

workbook = load_workbook(filename="Dataset.xlsx")
sheet = workbook['metastatic and primary CDR3s b']
sheet2 = workbook['DNAH9']
sheet3 = workbook['clinical']
sheet4 = workbook['PatientSurvivalChart']
sheet5 = workbook['Data']

for x in range(2, sheet2.max_row + 1):
    #Grab Patient ID
    Patient_ID = sheet2["A{cellRow}".format(cellRow = x)].value
    #Gets Amino Acid Changes
    Mutation = sheet2["E{cellRow}".format(cellRow = x)].value
    Ref_Allele = Mutation[0]
    Tumor_Seq_Allele = Mutation[len(Mutation)]
    #Calculates LorR
    LorR = LeftorRight()
    #Calculates Change in Charge of AA
    AA_Charge_Δ = Calc_AA_Charge_Δ(Ref_Allele,Tumor_Seq_Allele)
    #Finds days to death
    MonthsLeft = MonthsLeftFunc(Patient_ID,sheet3)
    #Calculates NCPR
    Complimentary_Func(Patient_ID,AA_Charge_Δ,LorR)
    #Assigns all values to rows in new datasheet
    sheet5["A{cellRow}".format(cellRow = x)] = Patient_ID+" "+Mutation
    sheet5["E{cellRow}".format(cellRow = x)] = Ref_Allele
    sheet5["F{cellRow}".format(cellRow = x)] = Tumor_Seq_Allele
    sheet5["G{cellRow}".format(cellRow = x)] = AA_Charge_Δ
    sheet5["D{cellRow}".format(cellRow = x)] = NCPR
    sheet5["H{cellRow}".format(cellRow = x)] = NCPR_CS
    sheet5["I{cellRow}".format(cellRow = x)] = MonthsLeft
    sheet5["K{cellRow}".format(cellRow = x)] = MutationPosition
    sheet5["L{cellRow}".format(cellRow = x)] = LeftorRight(MutationPosition)
    
# Save the spreadsheet
workbook.save(filename="EditedDataset.xlsx")

# Reopens SpreadSheet to use saved Complimentary Values
workbook = load_workbook(filename="EditedDataset.xlsx")
sheet = workbook['metastatic and primary CDR3s b']
sheet4 = workbook['PatientSurvivalChart']

#Makes survival Chart
PercentAlive(sheet,sheet4)

#81 NonCom Patients
#64 Comp Patients
#35 CompR Patients
#23 CompL Patients

#Saves Workbook
workbook.save(filename="EditedDataset2.xlsx")
